from time import *

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options


def connect(link):
    responce = requests.get(link).text
    soup = BeautifulSoup(responce, "html.parser")
    return soup


def name_search(address):
    block = address.find("div", class_="gr-elem-name")
    check_jx = block.find('h1').text
    product_name = check_jx.strip()
    return product_name


def article_search(address):
    art_mod = list()
    block = address.find_all("dl", class_="product-item-detail-properties")
    for sub_block in block:
        name = sub_block.find('dd').text.strip()
        art_mod.append(name)
    return art_mod[0]


def car_model_search(address):
    art_mod = list()
    block = address.find_all("dl", class_="product-item-detail-properties")
    for sub_block in block:
        name = sub_block.find('dd').text.strip()
        art_mod.append(name)
    return art_mod[1]


def price_search(address):
    block = address.find("div", class_="product-item-detail-price-current")
    price = block.text.strip()
    return price


def number_search(block, data_value):
    oe = block.find('div', {'data-value': data_value})
    numbers = oe.find_all('div', class_='gr-code-cell')
    result = []
    for number in numbers:
        result.append(number.text.strip())
    return result


def find_existing(address):
    existing = address.find('span', class_='gr-have')
    return existing.text.strip()


if __name__ == '__main__':
    stime = time()
    #Parsing all products
    with open('product_links.txt') as fr:
        product_base = fr.read().split('\n')

    for product in product_base:
        try:
            address = connect(product)
            sleep(3)
            name = name_search(address)
            article = article_search(address)
            car_model = car_model_search(address)
            price = price_search(address)
            try:
                original_numbers = number_search(address, 'OE')
                original_numbers = ', '.join(original_numbers)
            except:
                original_numbers = '-'
            try:
                analogue = number_search(address, 'CROSS_CODES')
                analogue = ', '.join(analogue)
            except:
                analogue = '-'
            have_check = find_existing(address)
            #Add the data to the list
            content = [name, article, car_model, price, original_numbers, analogue, have_check]
            #Add to the excel-table, check the type of product
            if product.find('rulevye_reyki') != -1:
                fn = 'reyki.xlsx'
                wb = load_workbook(fn)
                ws = wb['data']
                ws.append(content)
                wb.save(fn)
            if product.find('nasosy_gur') != -1:
                fb = 'pumps.xlsx'
                wf = load_workbook(fb)
                wk = wf['data']
                wk.append(content)
                wf.save(fb)
            if product.find('remkomplekty') != -1:
                ab = 'remcoms.xlsx'
                bc = load_workbook(ab)
                sw = bc['data']
                sw.append(content)
                bc.save(ab)
            #Print info for test
            print(name)
            print(article)
            print(car_model)
            print(price)
            print(original_numbers)
            print(analogue)
            print(have_check)
            print('#' * 10)
            print(' ' * 10)

        except:
            print(f'Link {product} is bad!')
            print('-' * 10)
            product += ' \n'
            with open('wrong_products.txt', 'a')  as file:
                file.write(product)
    etime = time()
    #Get work time
    full_time = (etime - stime)/60
    print(f'Parsing is complete! Time spent: {full_time} minutes')