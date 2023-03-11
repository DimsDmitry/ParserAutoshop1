from openpyxl import load_workbook

fn = "example1.xlsx"
wb = load_workbook(fn)
ws = wb['data']
ws['A5'] = 'Привет, мир!'
wb.save(fn)
wb.close()
