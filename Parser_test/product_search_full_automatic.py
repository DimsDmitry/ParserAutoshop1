import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from time import sleep
from time import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options

# Initialize Firefox/Gecko WebDriver
driver = webdriver.Firefox()
firefox_options = Options()
#firefox_options.set_headless()

def connect(link):
    responce = requests.get(link).text
    soup = BeautifulSoup(responce, "html.parser")
    return soup

def bad_window_hide():
    try:
        bad_window1 = driver.find_elements(by=By.CLASS_NAME, value='bxmaker__geoip__city__line-question-btn-no js-bxmaker__geoip__city__line-question-btn-no')
        bad_window1.click()
    except:
        sleep(1)
    try:
        bad_window2 = driver.find_elements(by=By.CLASS_NAME, value='b24-widget-button-popup-btn-hide')
        bad_window2.click()
    except:
        sleep(1)

def connect(link):
    responce = requests.get(link).text
    soup = BeautifulSoup(responce, "html.parser")
    return soup

def name_search(address):
    block = address.find("div", class_="gr-elem-name")
    check_jx = block.find('h1').text
    product_name = check_jx.strip()
    return product_name


def article_search(address):
    art_mod = list()
    block = address.find_all("dl", class_="product-item-detail-properties")
    for sub_block in block:
        name = sub_block.find('dd').text.strip()
        art_mod.append(name)
    return art_mod[0]


def car_model_search(address):
    art_mod = list()
    block = address.find_all("dl", class_="product-item-detail-properties")
    for sub_block in block:
        name = sub_block.find('dd').text.strip()
        art_mod.append(name)
    return art_mod[1]

def price_search(address):
    block = address.find("div", class_="product-item-detail-price-current")
    price = block.text.strip()
    return price

def number_search(block, data_value):
    oe = block.find('div', {'data-value': data_value})
    numbers = oe.find_all('div', class_='gr-code-cell')
    result = []
    for number in numbers:
        result.append(number.text.strip())
    return result


def find_existing(address):
    existing = address.find('span', class_='gr-have')
    return existing.text.strip()


if __name__ == '__main__':
    url = 'https://gelzer.ru/'
    driver = webdriver.Firefox(executable_path="../Parser_True_Final/geckodriver.exe", options=firefox_options)
    driver.get(url)
    stime = time()
    #button_element = driver.find_element_by_link_text('Alfa Romeo ')
    #marks = driver.find_elements_by_class_name('gr-mark-cell__wrapper')
    #i = 8
    marks = driver.find_elements(by=By.CLASS_NAME, value='gr-mark-cell__wrapper')

    for i in range(len(marks)):
        marks = driver.find_elements(by=By.CLASS_NAME, value='gr-mark-cell__wrapper')
        sleep(10)
        marks[i].click()
        sleep(10)

        bad_window_hide()

        product = driver.find_elements(by = By.CSS_SELECTOR, value = '.gr-row-models > div')
        try:
            product[0].click()
        except:
            try:
                product[1].click()
                with open('../Parser_True_Final/error_log.txt', 'a') as file:
                    file.write(f'Ошибка на итерации номер {i}')
            except:
                with open('../Parser_True_Final/error_log.txt', 'a') as file:
                    file.write(f'Ошибка на итерации номер {i}')
                    continue
        bad_window_hide()
        sleep(7)
        marks = driver.find_elements(by = By.CLASS_NAME, value = 'gr-filter-checkbox-wrapper')
        k = 0
        for k in range(len(marks)):
            marks[k].click()

            sleep(3)
            address = connect(driver.current_url)
            for a in address.find_all("a", {"class": "gr-product-item__border-inner"}):
                with open('../Parser_True_Final/product_links3.txt') as fr:
                    content = fr.read()
                print('Найдена ссылка:', a['href'])
                new_link = 'https://gelzer.ru' + a['href']
                if new_link in content:
                    print('Уже есть!')
                    continue
                with open('../Parser_True_Final/product_links3.txt', 'a') as file:
                    file.write(f'{new_link}\n')
        driver.get(url)
        sleep(5)

    #then parsing
    # parsing after search
    with open('../Parser_True_Final/product_links.txt') as fr:
        product_base = fr.read().split('\n')

    for product in product_base:
        try:
            address = connect(product)
            # sleep(3 Flask wtforms)
            name = name_search(address)
            article = article_search(address)
            car_model = car_model_search(address)
            price = price_search(address)
            try:
                original_numbers = number_search(address, 'OE')
                original_numbers = ', '.join(original_numbers)
            except:
                original_numbers = '-'
            try:
                analogue = number_search(address, 'CROSS_CODES')
                analogue = ', '.join(analogue)
            except:
                analogue = '-'
            have_check = find_existing(address)
            # складываем все данные в список
            content = [name, article, car_model, price, original_numbers, analogue, have_check]
            # записываем в excel-таблицу, проверяем вид товара
            if product.find('rulevye_reyki') != -1:
                fn = 'reyki.xlsx'
                wb = load_workbook(fn)
                ws = wb['data']
                ws.append(content)
                wb.save(fn)
            if product.find('nasosy_gur') != -1:
                fb = 'pumps.xlsx'
                wf = load_workbook(fb)
                wk = wf['data']
                wk.append(content)
                wf.save(fb)
            if product.find('remkomplekty') != -1:
                ab = 'remcoms.xlsx'
                bc = load_workbook(ab)
                sw = bc['data']
                sw.append(content)
                bc.save(ab)
            # печатаем для проверки
            print(name)
            print(article)
            print(car_model)
            print(price)
            print(original_numbers)
            print(analogue)
            print(have_check)
            print('#' * 10)
            print(' ' * 10)

        except:
            print(f'Ссылка {product} недействительна')
            print('-' * 10)
            product += ' \n'
            with open('../Parser_True_Final/wrong_links.txt', 'a') as file:
                file.write(product)
    etime = time()
    full_time = (etime - stime) / 60
    print(f'Все ссылки обработаны за {full_time} минут')


